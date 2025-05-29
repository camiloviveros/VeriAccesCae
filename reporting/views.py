# reporting/views.py
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse, FileResponse
from django.db.models import Count, Q
from django.utils import timezone
from datetime import datetime, timedelta
import json
import csv
import io
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

from .models import Report, GeneratedReport, ReportSchedule
from .serializers import ReportSerializer, GeneratedReportSerializer, ReportScheduleSerializer
from access_control.models import AccessLog, Visitor
from security.models import SecurityIncident

class ReportViewSet(viewsets.ModelViewSet):
    queryset = Report.objects.all().order_by('-created_at')
    serializer_class = ReportSerializer
    permission_classes = [IsAuthenticated]
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def generate(self, request, pk=None):
        """Genera un reporte basado en la definición"""
        report = self.get_object()
        
        # Determinar fechas según el período
        today = timezone.now().date()
        if report.period == 'daily':
            period_start = today
            period_end = today
        elif report.period == 'weekly':
            period_start = today - timedelta(days=7)
            period_end = today
        elif report.period == 'monthly':
            period_start = today - timedelta(days=30)
            period_end = today
        else:  # custom
            period_start = request.data.get('period_start', today - timedelta(days=30))
            period_end = request.data.get('period_end', today)
        
        # Obtener formato del reporte
        format_type = request.data.get('format', 'json')
        
        # Generar datos según el tipo de reporte
        data = self._generate_report_data(report, period_start, period_end)
        
        # Formatear respuesta según el formato solicitado
        if format_type == 'json':
            return Response({
                'report': ReportSerializer(report).data,
                'period_start': period_start,
                'period_end': period_end,
                'data': data
            })
        elif format_type == 'csv':
            return self._generate_csv_response(report, data, period_start, period_end)
        elif format_type == 'xlsx':
            return self._generate_excel_response(report, data, period_start, period_end)
        
        return Response({'error': 'Formato no soportado'}, status=status.HTTP_400_BAD_REQUEST)
    
    def _generate_report_data(self, report, period_start, period_end):
        """Genera los datos del reporte según su tipo"""
        data = {}
        
        if report.report_type == 'access_logs':
            # Reporte de registros de acceso
            logs = AccessLog.objects.filter(
                timestamp__date__gte=period_start,
                timestamp__date__lte=period_end
            )
            
            data = {
                'total_accesses': logs.count(),
                'granted': logs.filter(status='granted').count(),
                'denied': logs.filter(status='denied').count(),
                'by_day': [],
                'top_users': []
            }
            
            # Accesos por día
            current_date = period_start
            while current_date <= period_end:
                day_logs = logs.filter(timestamp__date=current_date)
                data['by_day'].append({
                    'date': current_date.isoformat(),
                    'total': day_logs.count(),
                    'granted': day_logs.filter(status='granted').count(),
                    'denied': day_logs.filter(status='denied').count()
                })
                current_date += timedelta(days=1)
            
        elif report.report_type == 'visitors':
            # Reporte de estadísticas de visitantes
            visitors = Visitor.objects.filter(
                created_at__date__gte=period_start,
                created_at__date__lte=period_end
            )
            
            data = {
                'total_visitors': visitors.count(),
                'by_type': {
                    'regular': visitors.filter(visitor_type='regular').count(),
                    'business': visitors.filter(visitor_type='business').count(),
                    'temporary': visitors.filter(visitor_type='temporary').count()
                },
                'by_status': {
                    'pending': visitors.filter(status='pending').count(),
                    'approved': visitors.filter(status='approved').count(),
                    'inside': visitors.filter(status='inside').count(),
                    'outside': visitors.filter(status='outside').count(),
                    'denied': visitors.filter(status='denied').count()
                }
            }
        
        elif report.report_type == 'incidents':
            # Reporte de incidentes de seguridad
            incidents = SecurityIncident.objects.filter(
                created_at__date__gte=period_start,
                created_at__date__lte=period_end
            )
            
            data = {
                'total_incidents': incidents.count(),
                'by_severity': {
                    'critical': incidents.filter(severity='critical').count(),
                    'high': incidents.filter(severity='high').count(),
                    'medium': incidents.filter(severity='medium').count(),
                    'low': incidents.filter(severity='low').count()
                },
                'by_status': {
                    'new': incidents.filter(status='new').count(),
                    'in_progress': incidents.filter(status='in_progress').count(),
                    'resolved': incidents.filter(status='resolved').count(),
                    'closed': incidents.filter(status='closed').count()
                },
                'average_resolution_time': self._calculate_avg_resolution_time(incidents)
            }
        
        return data
    
    def _calculate_avg_resolution_time(self, incidents):
        """Calcula el tiempo promedio de resolución de incidentes"""
        resolved = incidents.filter(resolved_at__isnull=False)
        if not resolved.exists():
            return 0
        
        total_time = timedelta()
        for incident in resolved:
            total_time += incident.resolved_at - incident.created_at
        
        avg_time = total_time / resolved.count()
        return avg_time.total_seconds() / 3600  # Convertir a horas
    
    def _generate_csv_response(self, report, data, period_start, period_end):
        """Genera una respuesta CSV"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Escribir encabezados
        writer.writerow(['Reporte:', report.name])
        writer.writerow(['Período:', f'{period_start} a {period_end}'])
        writer.writerow([])
        
        # Escribir datos según el tipo
        if report.report_type == 'access_logs':
            writer.writerow(['Métrica', 'Valor'])
            writer.writerow(['Total de accesos', data['total_accesses']])
            writer.writerow(['Accesos concedidos', data['granted']])
            writer.writerow(['Accesos denegados', data['denied']])
            writer.writerow([])
            writer.writerow(['Fecha', 'Total', 'Concedidos', 'Denegados'])
            for day in data['by_day']:
                writer.writerow([day['date'], day['total'], day['granted'], day['denied']])
        
        elif report.report_type == 'visitors':
            writer.writerow(['Métrica', 'Valor'])
            writer.writerow(['Total de visitantes', data['total_visitors']])
            writer.writerow([])
            writer.writerow(['Tipo', 'Cantidad'])
            for tipo, cantidad in data['by_type'].items():
                writer.writerow([tipo, cantidad])
            writer.writerow([])
            writer.writerow(['Estado', 'Cantidad'])
            for estado, cantidad in data['by_status'].items():
                writer.writerow([estado, cantidad])
        
        output.seek(0)
        response = HttpResponse(output, content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{report.name}_{period_start}_{period_end}.csv"'
        return response
    
    def _generate_excel_response(self, report, data, period_start, period_end):
        """Genera una respuesta Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"
        
        # Encabezados
        ws['A1'] = f'Reporte: {report.name}'
        ws['A2'] = f'Período: {period_start} a {period_end}'
        
        # Datos según el tipo
        if report.report_type == 'access_logs':
            ws['A4'] = 'Resumen'
            ws['A5'] = 'Total de accesos'
            ws['B5'] = data['total_accesses']
            ws['A6'] = 'Accesos concedidos'
            ws['B6'] = data['granted']
            ws['A7'] = 'Accesos denegados'
            ws['B7'] = data['denied']
            
            ws['A9'] = 'Detalle por día'
            ws['A10'] = 'Fecha'
            ws['B10'] = 'Total'
            ws['C10'] = 'Concedidos'
            ws['D10'] = 'Denegados'
            
            row = 11
            for day in data['by_day']:
                ws[f'A{row}'] = day['date']
                ws[f'B{row}'] = day['total']
                ws[f'C{row}'] = day['granted']
                ws[f'D{row}'] = day['denied']
                row += 1
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = FileResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{report.name}_{period_start}_{period_end}.xlsx"'
        return response

class GeneratedReportViewSet(viewsets.ModelViewSet):
    queryset = GeneratedReport.objects.all().order_by('-generated_at')
    serializer_class = GeneratedReportSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        # Filtrar por usuario si no es admin
        if not self.request.user.is_staff:
            queryset = queryset.filter(generated_by=self.request.user)
        return queryset

class ReportScheduleViewSet(viewsets.ModelViewSet):
    queryset = ReportSchedule.objects.all()
    serializer_class = ReportScheduleSerializer
    permission_classes = [IsAuthenticated]
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)